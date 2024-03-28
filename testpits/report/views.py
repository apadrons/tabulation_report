from django.shortcuts import render,get_object_or_404,HttpResponse
from .models import Project, Report
from .forms import ReportForm,ProjectForm
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from django.conf import settings
import os



def excel_convert(project,report):

    base_dir = settings.MEDIA_ROOT
    my_file = os.path.join(base_dir,'template.xlsx')
    wb = load_workbook(my_file)

    ws = wb['report']
    if report.before_excavation:
        img_path = os.path.join(base_dir,report.before_excavation.url[7:])
        img_before_excavation = Image(img_path)
        img_before_excavation.width = 264
        img_before_excavation.height = 211
        ws.add_image(img_before_excavation,'A60')
    if report.excavated_area:
        img_path = os.path.join(base_dir,report.excavated_area.url[7:])
        img_excavated_area = Image(img_path)
        img_excavated_area.width = 290
        img_excavated_area.height = 211
        ws.add_image(img_excavated_area,'B60')

    if report.after_excavation:
        img_path = os.path.join(base_dir,report.after_excavation.url[7:])
        img_after_excavation = Image(img_path)
        img_after_excavation.width = 285
        img_after_excavation.height = 211
        ws.add_image(img_after_excavation,'D60')
    
    ws['B3'].value = project.project_number
    ws['B4'].value = project.project_title
    ws['B5'].value = project.address

    ws['B8'].value = project.client
    ws['B9'].value = project.agreement_number
    ws['B10'].value = project.client_project_number

    ws['B12'].value = report.test_pit
    ws['B13'].value = report.field_coordinator
    ws['B14'].value = report.tech_name   
    ws['B15'].value = report.weather
    ws['B16'].value = report.date_project
    ws['B17'].value = report.timeproject

    ws['B24'].value = report.pavement_type
    ws['B25'].value = report.pavement_thickness

    ws['B29'].value = report.excavation_method
    ws['B30'].value = report.utility_size
    ws['B31'].value = report.material
    ws['B32'].value = report.utility_type
    ws['B33'].value = report.coordinate_system
    ws['B34'].value = report.field_coordination
    


    #Missing traffic control protocols, field observations and final remarks...

    new_file = os.path.join(base_dir,'converted.xlsx')
    wb.save(new_file)


# Create your views here.
def home_view(request):
  return render(request, 'report/home.html')

def projects_view(request):
    projects = Project.objects.all()
    context = {
        'projects' :projects
    }
    return render(request,'report/project_list.html',context)

def reports_view(request):
    reports = Report.objects.all()
    context = {
       'reports' : reports
    }
    return render(request,'report/report_list.html',context)

def single_report(request,test_pit):
    report = get_object_or_404(Report,test_pit = test_pit)
    context = {'report' : report}
    return render(request,'report/single.html',context)

def project_reports(request,project_number):
    project = Project.objects.get(project_number = project_number)
    reports = project.reports.all()
    print(reports)
    context = {'reports' : reports}
    return render(request,'report/project_reports.html',context)

def excel_convertion(request,test_pit):
    report = get_object_or_404(Report,test_pit = test_pit)
    project = get_object_or_404(Project,project_number = report.project_number)
    excel_convert(project,report)

    base_dir = settings.MEDIA_ROOT
    new_file = os.path.join(base_dir,'converted.xlsx')
#DOWNLOAD THE EXCEL DOCUMENT
    with open(new_file,'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="test_pit_{test_pit}.xlsx"'
    return response

def projects_form(request):
    context = {}
    form = ProjectForm(request.POST) 
    context['form'] = form
    if request.method == 'POST':
        if form.is_valid():  
            form.save()
            return projects_view(request)
        else:
                # The form is not valid.
                for field, error in form.errors.items():
                    context['error'] = error
                    print(error)
                return render(request, 'report/project.html', context)
        
    
    return render(request, 'report/project.html', context)

def report_form(request):
  context = {}
  form = ReportForm(request.POST,request.FILES) 
  context['form'] = form
  if request.method == 'POST':
    
    if form.is_valid():
        print('It is working')    
        form.save()
        return reports_view(request)
    else:
            # The form is not valid.
            for field, error in form.errors.items():
                context['error'] = error
            return render(request, 'report/report.html', context)
    
   
  return render(request, 'report/report.html', context)
