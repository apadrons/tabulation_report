from openpyxl import Workbook,load_workbook
from openpyxl.drawing.image import Image

def excel_convert(project,report):

    wb = load_workbook('./convertor/template.xlsx')

    ws = wb['report']

    #PROJECT INFO
    '''
    project_num = '0000'
    project_title = 'title'
    address = 'address'
    consultant = 'consultant'
    client = 'client'
    agreement = 'agreement'
    client_project_num = 'project_num'

    #Test Pit infor
    test_pit_num = 1
    field_coordinator = 'coord'
    tabulation_preparator = 'prepared'
    weather = 'sunny'
    date = '03/25/2024'
    time = '8:00 am'

    #Pavement information
    pavement_type = 'pavement type'
    pavement_thick = '4 in'

    #utility information
    excavation_method = 'Air Knife'
    utility_size = 'utility size'
    material = 'material'
    utility_type = 'Communications'
    coord_sys = 'NAD83 New Jersey State Planes - US foot'
    field_coordination = 'Regular Traffic Condition'
    '''

    ws['B3'].value = project.project_number
    ws['B4'].value = project.project_title
    ws['B5'].value = project.address

    ws['B8'].value = project.client
    ws['B9'].value = project.agreement_number
    ws['B10'].value = project.client_project_number

    ws['B12'].value = report.test_pit
    ws['B13'].value = report.field_coordinator
    ws['B14'].value = report.tech_name   
    ws['B15'].value = report.weather
    ws['B16'].value = report.date
    ws['B17'].value = report.time

    ws['B24'].value = report.pavement_type
    ws['B25'].value = report.pavement_thickness

    ws['B29'].value = report.excavation_method
    ws['B30'].value = report.utility_size
    ws['B31'].value = report.material
    ws['B32'].value = report.utility_type
    ws['B33'].value = report.coord_system
    ws['B34'].value = report.field_coordination

    #Missing traffic control protocols, field observations and final remarks...


    wb.save('sample.xlsx')


